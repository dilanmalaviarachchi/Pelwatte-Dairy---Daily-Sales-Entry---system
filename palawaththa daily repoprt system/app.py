import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime
from io import BytesIO
import os

def load_template(file_path):
    """Load the Excel template while preserving all formatting"""
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # Store headers (rows 1-19)
    headers = []
    for row in sheet.iter_rows(min_row=1, max_row=19, values_only=True):
        headers.append(row)
    
    # Get existing data (rows 20-51)
    existing_data = []
    for row in sheet.iter_rows(min_row=20, max_row=51, values_only=True):
        if any(cell is not None for cell in row):  # Only include rows with data
            # Convert Excel date numbers to datetime objects if needed
            if isinstance(row[0], (int, float)) and row[0] > 0:
                date_value = openpyxl.utils.datetime.from_excel(row[0])
                existing_data.append([date_value] + list(row[1:]))
            else:
                existing_data.append(row)
    
    return wb, headers, existing_data

def save_to_template(wb, headers, data, output_path):
    """Save data back to the template with original formatting"""
    sheet = wb.active
    
    # Clear existing data rows (20-51)
    for row in sheet.iter_rows(min_row=20, max_row=51):
        for cell in row:
            cell.value = None
    
    # Write headers (rows 1-19)
    for row_idx, row_data in enumerate(headers, 1):
        for col_idx, value in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Write data (rows 20-51)
    for row_idx, row_data in enumerate(data, 20):
        for col_idx, value in enumerate(row_data, 1):
            if col_idx == 1 and isinstance(value, datetime):  # Date column
                sheet.cell(row=row_idx, column=col_idx, value=value)
                sheet.cell(row=row_idx, column=col_idx).number_format = 'yyyy-mm-dd'
            else:
                sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Save the workbook
    wb.save(output_path)

def main():
    st.title("Pelwatte Dairy - Daily Sales Entry")
    st.subheader("Wattala RD Vehicle Allowance 2025")
    
    # Initialize session state
    if 'template_data' not in st.session_state:
        st.session_state.template_data = None
        st.session_state.template_headers = None
        st.session_state.workbook = None
        st.session_state.existing_data = []
        st.session_state.template_path = None
    
    # File upload section
    st.sidebar.header("1. Load Template")
    uploaded_file = st.sidebar.file_uploader("Upload your Excel template", type=["xlsx"])
    
    if uploaded_file is not None:
        if st.session_state.template_path != uploaded_file.name:
            try:
                # Save uploaded file temporarily
                with open("temp_template.xlsx", "wb") as f:
                    f.write(uploaded_file.getbuffer())
                
                # Load template
                wb, headers, existing_data = load_template("temp_template.xlsx")
                
                # Store in session state
                st.session_state.workbook = wb
                st.session_state.template_headers = headers
                st.session_state.existing_data = existing_data
                st.session_state.template_path = uploaded_file.name
                
                st.sidebar.success("Template loaded successfully!")
            except Exception as e:
                st.sidebar.error(f"Error loading template: {str(e)}")
    
   